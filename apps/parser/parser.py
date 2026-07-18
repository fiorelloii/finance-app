#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
parser.py

Parser PDF MPS -> Excel annuale, usando le POSIZIONI reali delle colonne nel PDF.

Regole principali:
- se il valore è sotto la colonna "Dare"  -> va in "uscita"
- se il valore è sotto la colonna "Avere" -> va in "entrata"
- i "Trasferimenti tra conti" verso/da BBVA restano nella tabella e nella
  sintesi per categoria, ma NON vengono conteggiati in:
    * Sintesi operazioni
    * Sintesi per mese

Input
- legge TUTTI i PDF presenti nella cartella ./resources
- il nome file del PDF contiene una data iniziale nel formato gg_mm_aaaa
  es. 02_04_2026_7971_Conto corrente...pdf
- il mese delle transazioni è il mese PRECEDENTE rispetto al mese nel nome file
  es. 04/2026 nel filename => transazioni di 03/2026

Output
- per ogni PDF aggiorna il file Excel dell'anno (<anno>.xlsx)
- il workbook contiene un solo foglio chiamato <anno>
- aggiunge in append solo le operazioni non ancora presenti
- ordina sempre la tabella per data/valuta/descrizione operazione
- rigenera ad ogni aggiornamento:
  * Sintesi operazioni (su tutti i dati, ESCLUSI i trasferimenti BBVA)
  * Sintesi per categoria (su TUTTI i dati)
  * Sintesi per mese (su tutti i dati, ESCLUSI i trasferimenti BBVA)

Colonne tabella
    data | valuta | descrizione operazione | uscita | entrata | categoria
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RESOURCES_DIR_NAME = "resources"
TABLE_COLUMNS = ['data', 'valuta', 'descrizione operazione', 'uscita', 'entrata', 'categoria']

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN = Side(border_style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}$')
AMOUNT_RE = re.compile(r'^\d{1,3}(?:\.\d{3})*,\d{2}$')

OWN_PARTY_PATTERNS = [
    'DOMENICOFIORILLI', 'DOMENICOBBVA', 'BBVAITM2', 'GIROCONTO', 'CONTIPERSONALI'
]


def normalize_spaces(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def compact(text: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", (text or "").upper())


def should_skip_generali_spesa(description: str, uscita: Optional[float], entrata: Optional[float]) -> bool:
    if not description:
        return False

    desc = compact(description)
    if 'GENERALI' not in desc:
        return False

    if uscita is None or pd.isna(uscita):
        return False

    uscita_value = float(uscita)
    if uscita_value <= 0:
        return False

    if entrata is not None and not pd.isna(entrata) and float(entrata) > 0:
        return False

    return abs(uscita_value - 75.5) <= 2


def amount_to_float(amount_txt: str) -> float:
    return float(amount_txt.replace(".", "").replace(",", "."))


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 11), 100)


def style_table(ws, start_row: int, start_col: int, nrows: int, ncols: int, header: bool = True) -> None:
    for r in range(start_row, start_row + nrows):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if r == start_row and header:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)


def resources_dir() -> Path:
    here = Path(__file__).resolve().parent
    return here / RESOURCES_DIR_NAME


def resolve_input_pdfs() -> List[Path]:
    base_dir = resources_dir()
    base_dir.mkdir(exist_ok=True)
    pdfs = sorted(p for p in base_dir.glob('*.pdf') if p.is_file() and not p.name.startswith('~$'))
    if pdfs:
        return pdfs
    here = Path(__file__).resolve().parent
    return sorted(p for p in here.glob('*.pdf') if p.is_file() and not p.name.startswith('~$'))


def year_month_from_filename(pdf_path: Path) -> Optional[Tuple[int, int]]:
    m = re.match(r'^(\d{2})_(\d{2})_(\d{4})\b', pdf_path.stem)
    if not m:
        return None
    month = int(m.group(2))
    year = int(m.group(3))
    prev_month = month - 1
    prev_year = year
    if prev_month == 0:
        prev_month = 12
        prev_year -= 1
    return prev_year, prev_month


def excel_row_key(rec: Dict) -> Tuple[str, str, str, str, str, str]:
    data = rec['data'].strftime('%Y-%m-%d') if pd.notna(rec['data']) else ''
    valuta = rec['valuta'].strftime('%Y-%m-%d') if pd.notna(rec['valuta']) else ''
    desc = normalize_spaces(str(rec['descrizione operazione']))
    uscita = '' if pd.isna(rec['uscita']) else f"{float(rec['uscita']):.2f}"
    entrata = '' if pd.isna(rec['entrata']) else f"{float(rec['entrata']):.2f}"
    categoria = '' if pd.isna(rec['categoria']) else str(rec['categoria'])
    return (data, valuta, desc, uscita, entrata, categoria)


def extract_page_fragments(page) -> List[Tuple[float, float, str]]:
    frags: List[Tuple[float, float, str]] = []

    def visitor(text, cm, tm, font_dict, font_size):
        t = text.strip()
        if not t:
            return
        x = float(tm[4])
        y = float(tm[5])
        frags.append((x, y, t))

    page.extract_text(visitor_text=visitor)
    return frags


def cluster_lines(frags: List[Tuple[float, float, str]], y_tol: float = 1.4) -> List[List[Tuple[float, float, str]]]:
    ordered = sorted(frags, key=lambda z: (-z[1], z[0]))
    lines: List[List[Tuple[float, float, str]]] = []
    current: List[Tuple[float, float, str]] = []
    current_y: Optional[float] = None
    for x, y, t in ordered:
        if current_y is None or abs(y - current_y) <= y_tol:
            current.append((x, y, t))
            current_y = y if current_y is None else (current_y + y) / 2
        else:
            lines.append(sorted(current, key=lambda z: z[0]))
            current = [(x, y, t)]
            current_y = y
    if current:
        lines.append(sorted(current, key=lambda z: z[0]))
    return lines


def detect_columns(lines: List[List[Tuple[float, float, str]]]) -> Optional[Dict[str, float]]:
    for line in lines:
        texts = [t.strip() for _, _, t in line]
        joined = ' '.join(texts)
        if 'Data' in joined and 'Valuta' in joined and 'Dare' in joined and 'Avere' in joined:
            cols = {}
            for x, _, t in line:
                ts = t.strip()
                if ts == 'Data':
                    cols['date_x'] = x
                elif 'Valuta' in ts:
                    cols['valuta_x'] = x
                elif 'Dare' in ts:
                    cols['dare_x'] = x
                elif 'Avere' in ts:
                    cols['avere_x'] = x
            if len(cols) == 4:
                cols['desc_x'] = cols['date_x'] + 30.0
                cols['dare_boundary'] = (cols['dare_x'] + cols['avere_x']) / 2.0
                return cols
    return None


def parse_transactions_from_pdf(pdf_path: Path) -> List[Dict]:
    reader = PdfReader(str(pdf_path))
    txs: List[Dict] = []

    for page in reader.pages:
        frags = extract_page_fragments(page)
        if not frags:
            continue
        lines = cluster_lines(frags)
        cols = detect_columns(lines)
        if not cols:
            continue

        in_table = False
        current = None

        for line in lines:
            texts_only = [t for _, _, t in line]
            line_text = normalize_spaces(' '.join(texts_only))
            if not line_text:
                continue
            if 'Data Descrizione Operazioni Valuta Dare Avere' in line_text or (
                'Data' in line_text and 'Valuta' in line_text and 'Dare' in line_text and 'Avere' in line_text
            ):
                in_table = True
                continue
            if not in_table:
                continue
            if line_text.startswith('Distinti saluti'):
                break
            if line_text.startswith('Pag.') or line_text.startswith('BANCAMONTEDEIPASCHI') or line_text.startswith('BANCA MONTE DEI PASCHI'):
                continue

            date_token = None
            valuta_token = None
            dare_token = None
            avere_token = None
            desc_parts = []

            for x, y, t in line:
                tt = normalize_spaces(t)
                if not tt:
                    continue
                if DATE_RE.match(tt) and x < cols['desc_x']:
                    date_token = tt
                elif DATE_RE.match(tt) and cols['valuta_x'] - 25 <= x <= cols['valuta_x'] + 40:
                    valuta_token = tt
                elif AMOUNT_RE.match(tt) and x >= cols['dare_x'] - 20 and x < cols['dare_boundary']:
                    dare_token = tt
                elif AMOUNT_RE.match(tt) and x >= cols['dare_boundary']:
                    avere_token = tt
                elif cols['desc_x'] <= x < cols['valuta_x'] - 5:
                    desc_parts.append(tt)

            is_new_tx = bool(date_token and valuta_token and (dare_token or avere_token))

            if is_new_tx:
                if current:
                    current['descrizione operazione'] = normalize_spaces(current['descrizione operazione'])
                    txs.append(current)
                current = {
                    'data': pd.to_datetime(date_token, format='%d/%m/%Y', errors='coerce'),
                    'valuta': pd.to_datetime(valuta_token, format='%d/%m/%Y', errors='coerce'),
                    'descrizione operazione': normalize_spaces(' '.join(desc_parts)),
                    'uscita': amount_to_float(dare_token) if dare_token else None,
                    'entrata': amount_to_float(avere_token) if avere_token else None,
                    'categoria': None,
                }
            else:
                if current and desc_parts:
                    extra = normalize_spaces(' '.join(desc_parts))
                    if extra:
                        current['descrizione operazione'] = normalize_spaces(
                            f"{current['descrizione operazione']} {extra}"
                        )

        if current:
            current['descrizione operazione'] = normalize_spaces(current['descrizione operazione'])
            txs.append(current)

    return txs


def categorize(description: str, direction: str) -> str:
    d = compact(description)
    if 'ACCREDITOEMOLUMENTI' in d or 'COMPETENZEMESE' in d or 'STIPENDIO' in d:
        return 'Stipendio'
    if 'BONIFICODALLESTERO' in d:
        return 'Rimborso/Assicurazione'
    if 'STORNOADDEBITODIRETTO' in d or 'RIMBORSO' in d:
        return 'Rimborso/Assicurazione' if ('ASSIC' in d or 'COVERGENIUS' in d) else 'Rimborso'
    if any(k in d for k in ['BBVAITM2', 'GIROCONTO', 'DOMENICOBBVA', 'DOMENICOFIORILLI']) and any(k in d for k in ['BONIFICOAVOSTROFAVORE', 'BONIST', 'BONSEPA', 'FILIALEDISPONENTE00560BONIST']):
        return 'Trasferimento tra conti'
    if 'RICEZIONEDENAROCONBANCOMATPAY' in d:
        return 'Entrata da terzi'
    if 'BONIFICOAVOSTROFAVORE' in d:
        return 'Entrata da terzi'
    if 'ORD' in d and not any(k in d for k in OWN_PARTY_PATTERNS) and 'AFAVORE' not in d and direction == 'entrata':
        return 'Entrata da terzi'
    if 'AFAVORE' in d and any(k in d for k in ['ELE', 'GIOVANNIGIAQUINTO', 'PIETRO', 'LUCAFIORILLI', 'SANTINOGALVAO']):
        return 'Bonifici a persone'
    if 'INVIODENAROCONBANCOMATPAY' in d:
        return 'Bonifici a persone'
    if any(k in d for k in ['DOLOMITIENERGIA', 'IREN', 'FASTWEB', 'WINDTRE', 'CONDOMINIO', 'ACQUA', 'FATTURAINTERNET', 'PAGAMENTORATADIMUTUO', 'MUTUO', 'ENERGIAELETTRICA']):
        return 'Casa'
    if any(k in d for k in ['GENERALI', 'DOMESTICANDGENERAL', 'SERVIZIDOMESTICANDGENERAL']):
        return 'Assicurazioni'
    if any(k in d for k in ['IMPOSTADIBOLLO', 'REGIONECAMPANIA', 'BOLLO']):
        return 'Tasse e imposte'
    if any(k in d for k in ['SOTTOSCRIZIONETITOLI', 'FONDICOMUNI', 'OICVM', 'ANIMA']):
        return 'Investimenti'
    if any(k in d for k in ['AMERICANEXPRESS', 'CARTAMONTEPASCHI', 'WIDIBA']):
        return 'Carte di credito'
    if any(k in d for k in ['ENI', 'ESSO', 'Q8', 'ITALO', 'ATAC', 'TAPGO', 'SERVICENAVIGO', 'FREENOW', 'APTNAPOLI', 'DISTRIBUTORE', 'SUPERGARAGE', 'ROTRANS', 'PV8004', 'CARBURANTI', 'ASPIT', 'CAIANELLO']):
        return 'Trasporti'
    if any(k in d for k in ['LUFTHANSA', 'TRANSAVIA', 'AEROPORTO', 'RELAY', 'ORLY', 'FRANKFURT']):
        return 'Viaggi'
    if any(k in d for k in ['SOLE365', 'EUROSPIN', 'LIDL', 'CONAD', 'SIGMA', 'MDBATTIPAGLIA', 'PICKUPSALERNOGELSO', 'FOORBAN']):
        return 'Spesa'
    if any(k in d for k in ['ZEROHEALTHYBAR', 'AGAVE', 'MCDONALD', 'DELIVEROO', 'BLACKROSES', 'OASI18', 'POORMANGER', 'ALBERTOMARCHETTI', 'LECINQUEPORTE', 'KINGSCROSS', 'LABOULANGERIE', 'KARTHY', 'CAFFEINA', 'ATTIMI', 'SIRIODUE', 'MAMMINA', 'CAFE', 'BAR', 'CENA', 'PRANZO', 'PANINO']):
        return 'Ristoranti e bar'
    if any(k in d for k in ['AMAZON', 'AMZN', 'PEPCO', 'DECATHLON', 'SCARPE', 'SHOPPINGCASA', 'LEROYMERLIN', 'NEXTSRL', 'LUSH', 'TIGOTA', 'TRONGONESPORT', 'MABOOK', 'HONDAMAGAZINE', 'CREAINFORMATICA', 'PAYPAL', 'BMW', 'SHOPPING']):
        return 'Shopping'
    if any(k in d for k in ['FARMACIE', 'FARMACIA', 'ANALISICLINICHE', 'ARTEMIS']):
        return 'Salute'
    if any(k in d for k in ['PETSHOUSE', 'FARMVET', 'VETERIN', 'GATTI']):
        return 'Animali'
    if 'PRELIEVODICONTANTE' in d:
        return 'Prelievo contante'
    return 'Entrata da terzi' if direction == 'entrata' else 'Uscite varie'


def is_bbva_internal_transfer(df: pd.DataFrame) -> pd.Series:
    desc = df['descrizione operazione'].fillna('').astype(str).str.upper()
    cat = df['categoria'].fillna('').astype(str)
    return cat.eq('Trasferimento tra conti') & desc.str.contains('BBVA', na=False)


def fallback_year_from_df(df: pd.DataFrame) -> int:
    years = df['data'].dropna().dt.year.unique().tolist()
    if not years:
        raise ValueError("Impossibile determinare l'anno.")
    return int(years[0])


def build_summary_tables(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    dfx = df.copy()
    dfx['uscita_tmp'] = dfx['uscita'].fillna(0.0)
    dfx['entrata_tmp'] = dfx['entrata'].fillna(0.0)

    # ESCLUDE dai riepiloghi operativi e mensili i trasferimenti interni verso/da BBVA
    dfx_no_bbva_transfers = dfx.loc[~is_bbva_internal_transfer(dfx)].copy()

    sintesi = pd.DataFrame(
        [
            ['Totale entrate', dfx_no_bbva_transfers['entrata_tmp'].sum()],
            ['Totale uscite', dfx_no_bbva_transfers['uscita_tmp'].sum()],
            ['Saldo netto', dfx_no_bbva_transfers['entrata_tmp'].sum() - dfx_no_bbva_transfers['uscita_tmp'].sum()],
            ['Numero operazioni', float(len(dfx_no_bbva_transfers))],
            ['Prima operazione', dfx_no_bbva_transfers['data'].min().strftime('%d/%m/%Y') if not dfx_no_bbva_transfers.empty else ''],
            ['Ultima operazione', dfx_no_bbva_transfers['data'].max().strftime('%d/%m/%Y') if not dfx_no_bbva_transfers.empty else ''],
        ],
        columns=['Voce', 'Valore'],
    )

    # Sintesi per categoria su TUTTI i dati, inclusi i trasferimenti BBVA
    per_categoria = (
        dfx.groupby('categoria', dropna=False)[['uscita_tmp', 'entrata_tmp']]
        .sum()
        .rename(columns={'uscita_tmp': 'uscite', 'entrata_tmp': 'entrate'})
        .sort_values(['uscite', 'entrate'], ascending=False)
        .reset_index()
    )

    # Sintesi per mese escludendo i trasferimenti BBVA
    dfx_no_bbva_transfers['mese'] = dfx_no_bbva_transfers['data'].dt.to_period('M').astype(str)
    per_mese = (
        dfx_no_bbva_transfers.groupby('mese', dropna=False)[['uscita_tmp', 'entrata_tmp']]
        .sum()
        .rename(columns={'uscita_tmp': 'uscite', 'entrata_tmp': 'entrate'})
        .reset_index()
        .sort_values('mese')
        .reset_index(drop=True)
    )

    return sintesi, per_categoria, per_mese


def read_existing_year_sheet(output_path: Path, sheet_name: str) -> pd.DataFrame:
    if not output_path.exists():
        return pd.DataFrame(columns=TABLE_COLUMNS)
    wb = load_workbook(output_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame(columns=TABLE_COLUMNS)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=TABLE_COLUMNS)
    header = [str(v).strip().lower() if v is not None else '' for v in rows[0]]
    if header[:len(TABLE_COLUMNS)] != TABLE_COLUMNS:
        return pd.DataFrame(columns=TABLE_COLUMNS)
    data_rows = []
    for row in rows[1:]:
        first_six = row[:6]
        if all(v is None for v in first_six):
            break
        data_rows.append(first_six)
    if not data_rows:
        return pd.DataFrame(columns=TABLE_COLUMNS)
    df = pd.DataFrame(data_rows, columns=TABLE_COLUMNS)
    df['data'] = pd.to_datetime(df['data'], dayfirst=True, errors='coerce')
    df['valuta'] = pd.to_datetime(df['valuta'], dayfirst=True, errors='coerce')
    df['uscita'] = pd.to_numeric(df['uscita'], errors='coerce')
    df['entrata'] = pd.to_numeric(df['entrata'], errors='coerce')
    df['descrizione operazione'] = df['descrizione operazione'].astype(str).map(normalize_spaces)
    df['categoria'] = df['categoria'].astype(str)
    return df


def merge_transactions(existing_df: pd.DataFrame, new_df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    existing_records = existing_df.to_dict('records') if not existing_df.empty else []
    existing_keys = {excel_row_key(rec) for rec in existing_records}
    appended = 0
    merged_records = list(existing_records)
    for rec in new_df.to_dict('records'):
        key = excel_row_key(rec)
        if key not in existing_keys:
            merged_records.append(rec)
            existing_keys.add(key)
            appended += 1
    merged_df = pd.DataFrame(merged_records, columns=TABLE_COLUMNS)
    if not merged_df.empty:
        merged_df['data'] = pd.to_datetime(merged_df['data'], errors='coerce')
        merged_df['valuta'] = pd.to_datetime(merged_df['valuta'], errors='coerce')
        merged_df['uscita'] = pd.to_numeric(merged_df['uscita'], errors='coerce')
        merged_df['entrata'] = pd.to_numeric(merged_df['entrata'], errors='coerce')
        merged_df = merged_df.sort_values(['data', 'valuta', 'descrizione operazione']).reset_index(drop=True)
    return merged_df, appended


def write_year_workbook(df: pd.DataFrame, output_path: Path, sheet_name: str) -> None:
    out = df.copy()
    out['data'] = out['data'].dt.strftime('%d/%m/%Y')
    out['valuta'] = out['valuta'].dt.strftime('%d/%m/%Y')
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer:
        out.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0, startcol=0)
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(out) + 1}'
    style_table(ws, 1, 1, len(out) + 1, 6, header=True)
    for row in range(2, len(out) + 2):
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '#,##0.00'
    sintesi, per_categoria, per_mese = build_summary_tables(df)
    start_col = 8
    ws.cell(row=1, column=start_col, value='Sintesi operazioni')
    ws.cell(row=1, column=start_col).fill = SECTION_FILL
    ws.cell(row=1, column=start_col).font = Font(bold=True)
    for i, row_values in enumerate([list(sintesi.columns)] + sintesi.values.tolist(), start=2):
        for j, value in enumerate(row_values, start=start_col):
            ws.cell(row=i, column=j, value=value)
    style_table(ws, 2, start_col, len(sintesi) + 1, 2, header=True)
    for r in range(3, 3 + len(sintesi)):
        if isinstance(ws.cell(r, start_col + 1).value, (int, float)):
            ws.cell(r, start_col + 1).number_format = '#,##0.00'
    cat_start = len(sintesi) + 5
    ws.cell(row=cat_start, column=start_col, value='Sintesi per categoria')
    ws.cell(row=cat_start, column=start_col).fill = SECTION_FILL
    ws.cell(row=cat_start, column=start_col).font = Font(bold=True)
    for i, row_values in enumerate([list(per_categoria.columns)] + per_categoria.values.tolist(), start=cat_start + 1):
        for j, value in enumerate(row_values, start=start_col):
            ws.cell(row=i, column=j, value=value)
    style_table(ws, cat_start + 1, start_col, len(per_categoria) + 1, 3, header=True)
    for r in range(cat_start + 2, cat_start + 2 + len(per_categoria)):
        ws.cell(r, start_col + 1).number_format = '#,##0.00'
        ws.cell(r, start_col + 2).number_format = '#,##0.00'
    mese_start = cat_start + len(per_categoria) + 4
    ws.cell(row=mese_start, column=start_col, value='Sintesi per mese')
    ws.cell(row=mese_start, column=start_col).fill = SECTION_FILL
    ws.cell(row=mese_start, column=start_col).font = Font(bold=True)
    for i, row_values in enumerate([list(per_mese.columns)] + per_mese.values.tolist(), start=mese_start + 1):
        for j, value in enumerate(row_values, start=start_col):
            ws.cell(row=i, column=j, value=value)
    style_table(ws, mese_start + 1, start_col, len(per_mese) + 1, 3, header=True)
    for r in range(mese_start + 2, mese_start + 2 + len(per_mese)):
        ws.cell(r, start_col + 1).number_format = '#,##0.00'
        ws.cell(r, start_col + 2).number_format = '#,##0.00'
    ws.column_dimensions['C'].width = 95
    autosize_columns(ws)
    wb.save(output_path)


def process_pdf(pdf_path: Path) -> Dict[str, object]:
    rows = parse_transactions_from_pdf(pdf_path)
    rows = [
        row for row in rows
        if not should_skip_generali_spesa(
            row.get('descrizione operazione'),
            row.get('uscita'),
            row.get('entrata'),
        )
    ]
    if not rows:
        raise ValueError(f'Nessun movimento trovato nel PDF: {pdf_path.name}')
    new_df = pd.DataFrame(rows, columns=TABLE_COLUMNS)
    new_df['categoria'] = new_df.apply(
        lambda r: categorize(
            r['descrizione operazione'],
            'entrata' if pd.notna(r['entrata']) and (pd.isna(r['uscita']) or float(r['entrata'] or 0) != 0) else 'uscita'
        ),
        axis=1,
    )
    new_df = new_df.sort_values(['data', 'valuta', 'descrizione operazione']).reset_index(drop=True)

    ym = year_month_from_filename(pdf_path)
    if ym:
        target_year, target_month = ym
    else:
        target_year = fallback_year_from_df(new_df)
        target_month = int(new_df['data'].dropna().dt.month.iloc[0]) if not new_df['data'].dropna().empty else 1
    sheet_name = str(target_year)
    output_path = pdf_path.parent / f'{sheet_name}.xlsx'
    existing_df = read_existing_year_sheet(output_path, sheet_name)
    merged_df, appended = merge_transactions(existing_df, new_df)
    write_year_workbook(merged_df, output_path, sheet_name)
    return {
        'pdf': pdf_path.name,
        'target_year': target_year,
        'target_month': target_month,
        'excel': output_path.name,
        'sheet': sheet_name,
        'pdf_rows': len(new_df),
        'appended': appended,
        'total_rows': len(merged_df),
    }


def main() -> None:
    pdfs = resolve_input_pdfs()
    if not pdfs:
        raise FileNotFoundError('Nessun PDF trovato nella cartella resources.')
    print(f'[INFO] Cartella resources: {resources_dir().resolve()}')
    print(f'[INFO] PDF trovati: {len(pdfs)}')
    results = []
    for pdf_path in pdfs:
        result = process_pdf(pdf_path)
        results.append(result)
        print(f"[OK] PDF: {result['pdf']}")
        print(f"[OK] Mese/anno dedotti dal filename: {result['target_month']:02d}/{result['target_year']}")
        print(f"[OK] File Excel: {result['excel']}")
        print(f"[OK] Foglio Excel: {result['sheet']}")
        print(f"[OK] Operazioni nel PDF: {result['pdf_rows']}")
        print(f"[OK] Operazioni nuove aggiunte: {result['appended']}")
        print(f"[OK] Operazioni totali nel foglio: {result['total_rows']}")
        print('---')
    print(f'[DONE] PDF elaborati: {len(results)}')


if __name__ == '__main__':
    main()